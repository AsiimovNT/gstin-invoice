"""
Professional Excel export for GST Invoice data.
"""

import io
from datetime import datetime
import openpyxl


def _fmt_inr(amount: float) -> str:
    """Indian number formatting: e.g. 520750 → ₹5,20,750"""
    amount = int(round(amount))
    s = str(abs(amount))
    if len(s) <= 3:
        result = s
    else:
        last3 = s[-3:]
        rest = s[:-3]
        groups = []
        while len(rest) > 2:
            groups.append(rest[-2:])
            rest = rest[:-2]
        if rest:
            groups.append(rest)
        result = ",".join(reversed(groups)) + "," + last3
    return f"₹{result}"
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Color palette ─────────────────────────────────────────────────────────────
HEADER_BG   = "1F3864"   # Deep navy
HEADER_FG   = "FFFFFF"   # White text
ALT_ROW_BG  = "EBF0FA"   # Light blue-grey
TOTAL_BG    = "D6E4F0"   # Slightly darker blue for totals row
BORDER_CLR  = "B0BEC5"   # Soft border


def _border(style="thin"):
    s = Side(style=style, color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()
    return cell


def _data_cell(ws, row, col, value, is_alt=False, align="left", number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", size=10)
    if is_alt:
        cell.fill = PatternFill("solid", fgColor=ALT_ROW_BG)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    cell.border = _border()
    if number_format:
        cell.number_format = number_format
    return cell


def write_excel(records: list[dict]) -> bytes:
    """
    Convert a list of invoice dicts to a formatted Excel workbook.
    Returns bytes ready to be written to a file or sent for download.
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Invoice Data ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Invoice Data"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"GST Invoice Summary  |  Extracted: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    title_cell.font = Font(name="Arial", bold=True, size=13, color=HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Blank separator
    ws.row_dimensions[2].height = 8

    # Headers (row 3)
    COLS = [
        "S.No.", "Supplier Name", "Supplier GSTIN",
        "Invoice No.", "Date", "Total Amount (₹)", "Source File"
    ]
    COL_WIDTHS = [7, 30, 20, 14, 14, 18, 30]

    for col_idx, (header, width) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        _header_cell(ws, 3, col_idx, header)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 36

    # Data rows
    total_amount = 0.0
    for row_offset, rec in enumerate(records, start=1):
        excel_row = 3 + row_offset
        is_alt = row_offset % 2 == 0

        amount_raw = rec.get("Total Amount (₹)", "—")
        try:
            amount_val = float(str(amount_raw).replace(",", ""))
            total_amount += amount_val
        except (ValueError, TypeError):
            amount_val = amount_raw

        _data_cell(ws, excel_row, 1, row_offset, is_alt, align="center")
        _data_cell(ws, excel_row, 2, rec.get("Supplier Name", "—"), is_alt)
        _data_cell(ws, excel_row, 3, rec.get("Supplier GSTIN", "—"), is_alt, align="center")
        _data_cell(ws, excel_row, 4, rec.get("Invoice No.", "—"), is_alt, align="center")
        _data_cell(ws, excel_row, 5, rec.get("Date", "—"), is_alt, align="center")

        # Amount column — numeric if parseable
        amt_cell = _data_cell(
            ws, excel_row, 6,
            amount_val if isinstance(amount_val, float) else amount_raw,
            is_alt, align="right",
            number_format='₹#,##0.00' if isinstance(amount_val, float) else None
        )

        _data_cell(ws, excel_row, 7, rec.get("Source File", "—"), is_alt)
        ws.row_dimensions[excel_row].height = 20

    # Totals row
    total_row = 3 + len(records) + 1
    ws.merge_cells(f"A{total_row}:E{total_row}")
    total_label = ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    total_label.font = Font(name="Arial", bold=True, size=11, color=HEADER_BG)
    total_label.fill = PatternFill("solid", fgColor=TOTAL_BG)
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    total_label.border = _border()

    total_val_cell = ws.cell(row=total_row, column=6, value=total_amount)
    total_val_cell.font = Font(name="Arial", bold=True, size=11, color=HEADER_BG)
    total_val_cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
    total_val_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_val_cell.border = _border()
    total_val_cell.number_format = '₹#,##0.00'

    # Filler in col G of total row
    filler = ws.cell(row=total_row, column=7, value="")
    filler.fill = PatternFill("solid", fgColor=TOTAL_BG)
    filler.border = _border()

    ws.row_dimensions[total_row].height = 24
    ws.freeze_panes = "A4"

    # ── Sheet 2: Summary Stats ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 22

    summary_headers = ["Metric", "Value"]
    for c, h in enumerate(summary_headers, 1):
        _header_cell(ws2, 1, c, h)

    amount_values = []
    for rec in records:
        try:
            amount_values.append(float(str(rec.get("Total Amount (₹)", "0")).replace(",", "")))
        except (ValueError, TypeError):
            pass

    stats = [
        ("Total Invoices",       len(records)),
        ("Total Amount (₹)",     _fmt_inr(total_amount)),
        ("Average Invoice (₹)",  _fmt_inr(sum(amount_values)/len(amount_values)) if amount_values else "—"),
        ("Highest Invoice (₹)",  _fmt_inr(max(amount_values)) if amount_values else "—"),
        ("Lowest Invoice (₹)",   _fmt_inr(min(amount_values)) if amount_values else "—"),
        ("Unique Suppliers",      len({r.get("Supplier Name") for r in records})),
        ("Extracted On",          datetime.now().strftime("%d-%b-%Y %H:%M")),
    ]

    for row_i, (metric, value) in enumerate(stats, start=2):
        is_alt = row_i % 2 == 0
        _data_cell(ws2, row_i, 1, metric, is_alt, align="left")
        _data_cell(ws2, row_i, 2, value, is_alt, align="right")
        ws2.row_dimensions[row_i].height = 22

    ws2.row_dimensions[1].height = 32

    # Return as bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def save_excel(records: list[dict], output_path: str) -> None:
    """Write the Excel file directly to disk."""
    data = write_excel(records)
    with open(output_path, "wb") as f:
        f.write(data)
